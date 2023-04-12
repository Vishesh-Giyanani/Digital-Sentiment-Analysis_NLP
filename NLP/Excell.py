import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('my_excel_file.xlsx')

# Select the sheet you want to modify
ws = wb['Sheet1']

# Define the column ranges for each category
categories = {'Appliances': ('A', 'F'), 'Locks': ('G', 'L'), 'Interio': ('M', 'R'), 'Security': ('S', 'X')}

# Set the colors for each category
colors = {'Appliances': 'FFA500', 'Locks': '87CEFA', 'Interio': 'FFFF00', 'Security': '00FF00'}

# Loop through each category and apply the appropriate color to the columns
for category, col_range in categories.items():
    start_col, end_col = col_range
    for col in ws.iter_cols(min_col=ws[start_col+'1'].column, max_col=ws[end_col+'1'].column):
        for cell in col:
            if cell.column_letter.startswith(start_col):
                cell.fill = openpyxl.styles.PatternFill(start_color=colors[category], end_color=colors[category], fill_type='solid')

# Create new sheets for each category
for category, col_range in categories.items():
    start_col, end_col = col_range
    ws_new = wb.create_sheet(category)
    ws_new.append([cell.value for cell in ws[start_col+'1':end_col+'1'][0]])
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row if start_col <= cell.column_letter <= end_col]
        if any(values):
            ws_new.append(values)

# Save the modified Excel file
wb.save('my_excel_file.xlsx')
