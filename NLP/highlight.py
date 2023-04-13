import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side,Font


# Read the input CSV file
with open('./NLP/final3.csv', 'r') as infile:
    reader = csv.reader(infile)
    data = [tuple(row) for row in reader]
  
# remove first row
with open('./NLP/final3.csv', 'r') as infile:
    reader = csv.reader(infile)
    data = [tuple(row) for row in reader][1:] 


# Create a new workbook and select the active worksheet
workbook = Workbook()
worksheet = workbook.active

# Define the column numbers you want to sort and the order to use
columns_to_sort = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27]
sort_order = {'Positive': 1, 'Neutral': 2, 'Negative': 3}

# Sort the data by the selected columns and sort order
data.sort(key=lambda row: [sort_order.get(row[col], 0) for col in columns_to_sort])

# Define the column numbers you want to highlight and the colors to use
columns_to_highlight = {1:'FFFFFF',2:'FFFFFF',3:'FFFFFF',4: 'FFFFFF',
                        5: 'FFC0CB', 6: 'FFC0CB', 7: 'FFC0CB', 8: 'FFC0CB', 9: 'FFC0CB',10: 'FFC0CB',
                        11: 'ADD8E6', 12: 'ADD8E6', 13: 'ADD8E6', 14: 'ADD8E6', 15: 'ADD8E6',16: 'ADD8E6',
                        17: '98FB98', 18: '98FB98', 19: '98FB98', 20: '98FB98', 21: '98FB98',22: '98FB98',
                        23: 'E6E6FA', 24: 'E6E6FA', 25: 'E6E6FA', 26: 'E6E6FA', 27: 'E6E6FA',28:'E6E6FA',
                        29:'FFFFFF', 30:'FFFFFF',31:'FFFFFF', 32:'FFFFFF', 33:'FFFFFF', 
                        34:'FFFFFF', 35:'FFFFFF', 36:'FFFFFF', 37:'FFFFFF', 38:'FFFFFF', }

# Iterate through the keys of columns_to_highlight and apply the fill style to the cells in each column
for col_num in columns_to_highlight.keys():
    color = columns_to_highlight[col_num]
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    for row_num, row in enumerate(data, start=1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = row[col_num - 1]
        cell.fill = fill
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'),
                                      right=openpyxl.styles.Side(border_style='thin', color='000000'),
                                      top=openpyxl.styles.Side(border_style='thin', color='000000'),
                                      bottom=openpyxl.styles.Side(border_style='thin', color='000000'))

# Define the fill style for the first row
header_fill = PatternFill(start_color='F5F5DC', end_color='F5F5DC', fill_type='solid')

# Define the font style for the first row
header_font = Font(bold=True)

# Highlight the first row with the fill style
for col in worksheet.columns:
    cell = col[0]
    cell.fill = header_fill
    cell.font = header_font

   
# Write the output to a new CSV file
with open('./NLP/final4.csv', 'w', newline='') as outfile:
    writer = csv.writer(outfile)
    for row in data:
        writer.writerow(row)

workbook.save('./NLP/final4.xlsx')